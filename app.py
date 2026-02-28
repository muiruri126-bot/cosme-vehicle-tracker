"""
COSME Project – Vehicle Booking & Movement Tracker  (v2 – full features)
=========================================================================
Main Flask application – routes, configuration, and database init.

Features added in v2:
  1. User authentication & roles (admin / driver / requester)
  2. Driver assignment to bookings
  3. Email notification on booking approval  (configurable – off by default)
  4. Vehicle maintenance scheduling
  5. Fuel cost tracking & budget report per project code
  6. Export trip report to Excel (.xlsx)
  7. Calendar view of bookings (FullCalendar.js)

How to run
----------
1.  pip install -r requirements.txt
2.  python app.py          # starts the dev server on http://127.0.0.1:5000

A default **admin** account is created on first run:
    username: admin   |   password: admin123
"""

import io
import os
from datetime import datetime, timedelta, timezone
from functools import wraps
from urllib.parse import urlsplit

from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_login import (
    LoginManager,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_mail import Mail, Message
from flask_wtf.csrf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models import (
    AuditLog,
    Booking,
    MaintenanceRecord,
    Trip,
    User,
    Vehicle,
    check_booking_conflict,
    db,
)

# ── App & config ─────────────────────────────────────────────────────────────

load_dotenv()  # load .env file if present

app = Flask(__name__)

basedir = os.path.abspath(os.path.dirname(__file__))
database_url = os.environ.get(
    "DATABASE_URL", "sqlite:///" + os.path.join(basedir, "tracker.db")
)
# Heroku / PythonAnywhere may provide postgres:// but SQLAlchemy 2.x needs postgresql://
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "cosme-dev-secret-key")

# ── Session timeout (15 minutes of inactivity) ──────────────────────────────
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=15)

# ── Mail config (disabled by default – set MAIL_ENABLED=1 env var to turn on)
app.config["MAIL_ENABLED"] = os.environ.get("MAIL_ENABLED", "0") == "1"
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME", "")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD", "")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get(
    "MAIL_DEFAULT_SENDER", "noreply@cosme-project.org"
)

db.init_app(app)
mail = Mail(app)
csrf = CSRFProtect(app)
limiter = Limiter(app=app, key_func=get_remote_address, default_limits=[])

# ── Pagination ───────────────────────────────────────────────────────────────
PER_PAGE = 20

# ── Flask-Login setup ────────────────────────────────────────────────────────

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message_category = "warning"


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


@app.before_request
def check_session_timeout():
    """Log out the user if they have been inactive for more than 15 minutes."""
    if current_user.is_authenticated:
        now = datetime.now(timezone.utc)
        last_active = session.get("last_active")
        if last_active is not None:
            # Ensure last_active is timezone-aware (UTC) for safe subtraction
            if hasattr(last_active, 'tzinfo') and last_active.tzinfo is None:
                last_active = last_active.replace(tzinfo=timezone.utc)
            elapsed = (now - last_active).total_seconds()
            if elapsed > 15 * 60:  # 15 minutes
                logout_user()
                session.clear()
                flash("Your session has expired due to inactivity. Please log in again.", "warning")
                return redirect(url_for("login"))
        session["last_active"] = now
        session.permanent = True


@app.before_request
def force_password_change():
    """Redirect users who must change their password before accessing any page."""
    if current_user.is_authenticated and getattr(current_user, "must_change_password", False):
        allowed_endpoints = ("change_password", "logout", "static")
        if request.endpoint not in allowed_endpoints:
            flash("You must change your password before continuing.", "warning")
            return redirect(url_for("change_password"))


# ── Role-based access decorator ─────────────────────────────────────────────


def role_required(*roles):
    """Decorator: require the current user to have one of the given roles."""

    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if current_user.role not in roles:
                flash("You do not have permission to access this page.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)

        return decorated_function

    return decorator


# ── Audit-log helper ─────────────────────────────────────────────────────────


def log_action(action, entity_type, entity_id=None, details=None):
    """Record an audit-log entry for the current user."""
    entry = AuditLog(
        user_id=current_user.id if current_user.is_authenticated else None,
        username=current_user.username if current_user.is_authenticated else "system",
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
    )
    db.session.add(entry)
    db.session.commit()


# ── Email helper ─────────────────────────────────────────────────────────────


def send_notification(subject, recipients, body):
    """Send an email notification (only if MAIL_ENABLED is true)."""
    if not app.config["MAIL_ENABLED"]:
        return  # silently skip
    try:
        msg = Message(subject=subject, recipients=recipients, body=body)
        mail.send(msg)
    except Exception as e:
        app.logger.error(f"Failed to send email: {e}")


# ── Create DB tables & default admin ─────────────────────────────────────────

with app.app_context():
    db.create_all()
    # Create default admin if none exists
    if not User.query.filter_by(username="admin").first():
        admin = User(
            username="admin",
            email="admin@cosme-project.org",
            full_name="System Admin",
            role="admin",
            must_change_password=True,
        )
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  AUTHENTICATION                                                         ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("10/minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        if not username or not password:
            flash("Please enter both username and password.", "danger")
            return render_template("auth/login.html")

        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            if not user.is_active_user:
                flash("Your account has been deactivated. Contact an administrator.", "danger")
                return render_template("auth/login.html")
            login_user(user)
            flash(f"Welcome back, {user.full_name}!", "success")
            next_page = request.args.get("next")
            # Prevent open-redirect: reject absolute URLs
            if next_page:
                parsed = urlsplit(next_page)
                if parsed.netloc or parsed.scheme:
                    next_page = None
            return redirect(next_page or url_for("dashboard"))
        flash("Invalid username or password.", "danger")

    return render_template("auth/login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    """Let a user change their own password (also used for forced change on first login)."""
    forced = current_user.must_change_password

    if request.method == "POST":
        current_pw = request.form.get("current_password", "")
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")

        errors = []
        if not current_user.check_password(current_pw):
            errors.append("Current password is incorrect.")
        if not new_pw or len(new_pw) < 6:
            errors.append("New password must be at least 6 characters.")
        if new_pw != confirm_pw:
            errors.append("New passwords do not match.")
        if current_pw and new_pw and current_pw == new_pw:
            errors.append("New password must be different from the current one.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("auth/change_password.html", forced=forced)

        current_user.set_password(new_pw)
        current_user.must_change_password = False
        db.session.commit()
        log_action("edit", "User", current_user.id, "Changed own password")
        flash("Your password has been changed successfully.", "success")
        return redirect(url_for("dashboard"))

    return render_template("auth/change_password.html", forced=forced)


@app.route("/register", methods=["GET", "POST"])
@limiter.limit("5/minute")
def register():
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        username = request.form.get("username", "").strip().lower()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        errors = []

        # Required fields
        if not full_name:
            errors.append("Full name is required.")
        if not username:
            errors.append("Username is required.")
        elif len(username) < 3:
            errors.append("Username must be at least 3 characters.")
        elif not username.replace("_", "").replace(".", "").isalnum():
            errors.append("Username may only contain letters, numbers, dots, and underscores.")
        if not email:
            errors.append("Email is required.")
        elif "@" not in email or "." not in email.split("@")[-1]:
            errors.append("Please enter a valid email address.")

        # Password checks
        if not password:
            errors.append("Password is required.")
        elif len(password) < 6:
            errors.append("Password must be at least 6 characters.")
        if password != password2:
            errors.append("Passwords do not match.")

        # Uniqueness
        if not errors:
            if User.query.filter_by(username=username).first():
                errors.append("Username already taken.")
            if User.query.filter_by(email=email).first():
                errors.append("Email already registered.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("auth/register.html")

        user = User(
            username=username,
            email=email,
            full_name=full_name,
            role="requester",  # default role; admin can change later
        )
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash("Account created! You can now log in.", "success")
        return redirect(url_for("login"))

    return render_template("auth/register.html")


# ── Password reset (self-service) ────────────────────────────────────────────


@app.route("/forgot-password", methods=["GET", "POST"])
@limiter.limit("5/minute")
def forgot_password():
    """Let a user request a password-reset email."""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        # Always show the same message to prevent email enumeration
        flash("If that email is registered, a reset link has been sent.", "info")

        user = User.query.filter_by(email=email).first()
        if user and user.is_active_user:
            token = user.generate_reset_token()
            db.session.commit()
            reset_url = url_for("reset_password", token=token, _external=True)
            send_notification(
                subject="Vehicle Request Tracker – Password Reset",
                recipients=[user.email],
                body=(
                    f"Hello {user.full_name},\n\n"
                    f"A password reset was requested for your account ({user.username}).\n\n"
                    f"Click the link below to reset your password (valid for 24 hours):\n"
                    f"{reset_url}\n\n"
                    f"If you did not request this, please ignore this email.\n\n"
                    f"— Vehicle Request Tracker"
                ),
            )
        return redirect(url_for("login"))

    return render_template("auth/forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
@limiter.limit("10/minute")
def reset_password(token):
    """Validate a password-reset token and let the user choose a new password."""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    user = User.verify_reset_token(token)
    if user is None:
        flash("Invalid or expired reset link. Please request a new one.", "danger")
        return redirect(url_for("forgot_password"))

    if request.method == "POST":
        new_pw = request.form.get("new_password", "")
        confirm_pw = request.form.get("confirm_password", "")

        errors = []
        if not new_pw or len(new_pw) < 6:
            errors.append("Password must be at least 6 characters.")
        if new_pw != confirm_pw:
            errors.append("Passwords do not match.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("auth/reset_password.html", token=token)

        user.set_password(new_pw)
        user.clear_reset_token()
        user.must_change_password = False
        db.session.commit()
        flash("Your password has been reset. Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("auth/reset_password.html", token=token)


# ── User management (admin only) ─────────────────────────────────────────────


@app.route("/users")
@role_required("admin")
def user_list():
    page = request.args.get("page", 1, type=int)
    pagination = User.query.order_by(User.full_name).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    return render_template("auth/user_list.html", users=pagination.items, pagination=pagination)


@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@role_required("admin")
def user_edit(user_id):
    user = db.get_or_404(User, user_id)
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        role = request.form.get("role", "")
        is_active = "is_active_user" in request.form

        errors = []

        if not full_name:
            errors.append("Full name is required.")
        if not email or "@" not in email:
            errors.append("A valid email is required.")

        # Check email uniqueness (exclude this user)
        existing = User.query.filter(User.email == email, User.id != user.id).first()
        if existing:
            errors.append(f"Email '{email}' is already used by {existing.username}.")

        # Valid role
        if role not in ("admin", "driver", "requester"):
            errors.append("Invalid role selected.")

        # Prevent admin from deactivating or demoting themselves
        if user.id == current_user.id:
            if not is_active:
                errors.append("You cannot deactivate your own account.")
            if role != "admin":
                errors.append("You cannot remove your own admin role.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("auth/user_edit.html", user=user)

        user.full_name = full_name
        user.email = email
        user.role = role
        user.is_active_user = is_active
        db.session.commit()
        log_action("edit", "User", user.id, f"Updated user '{user.username}': name={full_name}, email={email}, role={role}, active={is_active}")
        flash(f"User {user.username} updated.", "success")
        return redirect(url_for("user_list"))
    return render_template("auth/user_edit.html", user=user)


@app.route("/users/<int:user_id>/reset-password", methods=["POST"])
@role_required("admin")
def admin_reset_password(user_id):
    """Admin action: reset a user's password and force them to change it on next login."""
    user = db.get_or_404(User, user_id)
    new_pw = request.form.get("new_password", "").strip()
    if not new_pw or len(new_pw) < 6:
        flash("Password must be at least 6 characters.", "danger")
        return redirect(url_for("user_edit", user_id=user.id))
    user.set_password(new_pw)
    user.must_change_password = True
    db.session.commit()
    log_action("edit", "User", user.id, f"Admin reset password for '{user.username}' (forced change)")
    flash(f"Password for {user.username} has been reset. They will be asked to change it on next login.", "success")
    return redirect(url_for("user_edit", user_id=user.id))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@role_required("admin")
def user_delete(user_id):
    user = db.get_or_404(User, user_id)
    if user.id == current_user.id:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("user_list"))
    # Remove associated bookings & trips
    for booking in user.bookings_requested:
        if booking.trip:
            db.session.delete(booking.trip)
        db.session.delete(booking)
    # Unassign from bookings where user was driver
    for booking in user.bookings_driven:
        booking.driver_id = None
    username_deleted = user.username
    db.session.delete(user)
    db.session.commit()
    log_action("delete", "User", user_id, f"Deleted user '{username_deleted}' and associated records")
    flash(f"User '{username_deleted}' and their associated records have been deleted.", "success")
    return redirect(url_for("user_list"))


# ── User profile (self-service) ──────────────────────────────────────────────


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    """Let users view and edit their own profile (full name, email)."""
    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()

        errors = []
        if not full_name:
            errors.append("Full name is required.")
        if not email or "@" not in email or "." not in email.split("@")[-1]:
            errors.append("A valid email is required.")

        # Email uniqueness (exclude current user)
        existing = User.query.filter(User.email == email, User.id != current_user.id).first()
        if existing:
            errors.append(f"Email '{email}' is already used by another account.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("auth/profile.html")

        current_user.full_name = full_name
        current_user.email = email
        db.session.commit()
        log_action("edit", "User", current_user.id, f"Updated own profile: name={full_name}, email={email}")
        flash("Profile updated successfully.", "success")
        return redirect(url_for("profile"))

    return render_template("auth/profile.html")


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  DASHBOARD                                                              ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/")
@login_required
def dashboard():
    """Home page – quick stats and upcoming approved bookings."""
    vehicle_count = Vehicle.query.count()
    pending_count = Booking.query.filter_by(status="pending").count()
    maintenance_due = MaintenanceRecord.query.filter_by(status="scheduled").count()
    upcoming = (
        Booking.query.filter_by(status="approved")
        .order_by(Booking.start_datetime_planned)
        .all()
    )
    return render_template(
        "dashboard.html",
        vehicle_count=vehicle_count,
        pending_count=pending_count,
        maintenance_due=maintenance_due,
        upcoming=upcoming,
    )


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  VEHICLES                                                               ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/vehicles")
@login_required
def vehicle_list():
    page = request.args.get("page", 1, type=int)
    pagination = Vehicle.query.order_by(Vehicle.registration_number).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    return render_template("vehicles/list.html", vehicles=pagination.items, pagination=pagination)


@app.route("/vehicles/add", methods=["GET", "POST"])
@role_required("admin")
def vehicle_add():
    if request.method == "POST":
        reg = request.form.get("registration_number", "").strip().upper()
        make = request.form.get("make", "").strip()
        model = request.form.get("model", "").strip()
        status = request.form.get("status", "available")

        errors = []
        if not reg:
            errors.append("Registration number is required.")
        elif len(reg) < 3:
            errors.append("Registration number is too short.")
        if not make:
            errors.append("Vehicle make is required.")
        if not model:
            errors.append("Vehicle model is required.")
        if status not in ("available", "maintenance"):
            errors.append("Invalid vehicle status.")

        # Uniqueness
        if reg and Vehicle.query.filter_by(registration_number=reg).first():
            errors.append(f"Registration number '{reg}' already exists.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("vehicles/add.html")

        v = Vehicle(
            registration_number=reg,
            make=make,
            model=model,
            status=status,
        )
        db.session.add(v)
        db.session.commit()
        log_action("create", "Vehicle", v.id, f"Registered vehicle '{reg}' ({make} {model})")
        flash("Vehicle registered successfully.", "success")
        return redirect(url_for("vehicle_list"))
    return render_template("vehicles/add.html")


@app.route("/vehicles/<int:vehicle_id>/edit", methods=["GET", "POST"])
@role_required("admin")
def vehicle_edit(vehicle_id):
    vehicle = db.get_or_404(Vehicle, vehicle_id)
    if request.method == "POST":
        reg = request.form.get("registration_number", "").strip().upper()
        make = request.form.get("make", "").strip()
        model = request.form.get("model", "").strip()
        status = request.form.get("status", "")

        errors = []
        if not reg:
            errors.append("Registration number is required.")
        elif len(reg) < 3:
            errors.append("Registration number is too short.")
        if not make:
            errors.append("Vehicle make is required.")
        if not model:
            errors.append("Vehicle model is required.")
        if status not in ("available", "in_use", "maintenance"):
            errors.append("Invalid vehicle status.")

        # Uniqueness (exclude this vehicle)
        if reg:
            dup = Vehicle.query.filter(
                Vehicle.registration_number == reg, Vehicle.id != vehicle.id
            ).first()
            if dup:
                errors.append(f"Registration number '{reg}' is already used by another vehicle.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("vehicles/edit.html", vehicle=vehicle)

        vehicle.registration_number = reg
        vehicle.make = make
        vehicle.model = model
        vehicle.status = status
        db.session.commit()
        log_action("edit", "Vehicle", vehicle.id, f"Updated vehicle '{reg}': make={make}, model={model}, status={status}")
        flash("Vehicle updated.", "success")
        return redirect(url_for("vehicle_list"))
    return render_template("vehicles/edit.html", vehicle=vehicle)


@app.route("/vehicles/<int:vehicle_id>/delete", methods=["POST"])
@role_required("admin")
def vehicle_delete(vehicle_id):
    vehicle = db.get_or_404(Vehicle, vehicle_id)
    # Delete associated maintenance records
    for rec in vehicle.maintenance_records:
        db.session.delete(rec)
    # Delete associated bookings and their trips
    for booking in vehicle.bookings:
        if booking.trip:
            db.session.delete(booking.trip)
        db.session.delete(booking)
    db.session.delete(vehicle)
    db.session.commit()
    log_action("delete", "Vehicle", vehicle_id, f"Deleted vehicle '{vehicle.registration_number}' and all associated records")
    flash(f"Vehicle '{vehicle.registration_number}' and all associated records have been deleted.", "success")
    return redirect(url_for("vehicle_list"))


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  BOOKINGS                                                               ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/bookings")
@login_required
def booking_list():
    status_filter = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)
    query = Booking.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    pagination = query.order_by(Booking.start_datetime_planned.desc()).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    return render_template(
        "bookings/list.html",
        bookings=pagination.items,
        pagination=pagination,
        current_status=status_filter,
    )


@app.route("/bookings/add", methods=["GET", "POST"])
@login_required
def booking_add():
    vehicles = Vehicle.query.order_by(Vehicle.registration_number).all()
    drivers = User.query.filter_by(role="driver").order_by(User.full_name).all()

    if request.method == "POST":
        errors = []

        # ── Parse & validate required fields ──────────────────────────
        try:
            vehicle_id = int(request.form.get("vehicle_id", 0))
        except (ValueError, TypeError):
            vehicle_id = 0
        if not vehicle_id:
            errors.append("Please select a vehicle.")

        start_raw = request.form.get("start_datetime_planned", "")
        end_raw = request.form.get("end_datetime_planned", "")
        start_dt = end_dt = None
        if not start_raw:
            errors.append("Planned start date/time is required.")
        else:
            try:
                start_dt = datetime.fromisoformat(start_raw)
            except ValueError:
                errors.append("Invalid start date/time format.")

        if not end_raw:
            errors.append("Planned end date/time is required.")
        else:
            try:
                end_dt = datetime.fromisoformat(end_raw)
            except ValueError:
                errors.append("Invalid end date/time format.")

        route_from = request.form.get("route_from", "").strip()
        route_to = request.form.get("route_to", "").strip()
        purpose = request.form.get("purpose", "").strip()

        if not route_from:
            errors.append("Route From is required.")
        if not route_to:
            errors.append("Route To is required.")
        if not purpose:
            errors.append("Purpose is required.")

        # ── Date logic checks ─────────────────────────────────────────
        if start_dt and end_dt:
            if end_dt <= start_dt:
                errors.append("End date/time must be after start date/time.")
            if start_dt < datetime.now():
                errors.append("Start date/time cannot be in the past.")

        # ── Vehicle status check ──────────────────────────────────────
        if vehicle_id:
            veh = db.session.get(Vehicle, vehicle_id)
            if not veh:
                errors.append("Selected vehicle does not exist.")
            elif veh.status == "maintenance":
                errors.append(
                    f"Vehicle {veh.registration_number} is currently under maintenance "
                    f"and cannot be booked."
                )

        # ── Conflict check ────────────────────────────────────────────
        if not errors and start_dt and end_dt:
            conflict = check_booking_conflict(vehicle_id, start_dt, end_dt)
            if conflict:
                errors.append(
                    f"This vehicle is already booked between "
                    f"{conflict.start_datetime_planned.strftime('%Y-%m-%d %H:%M')} and "
                    f"{conflict.end_datetime_planned.strftime('%Y-%m-%d %H:%M')} "
                    f"(Booking #{conflict.id} by {conflict.requester_name})."
                )

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template(
                "bookings/add.html", vehicles=vehicles, drivers=drivers
            )

        driver_id = request.form.get("driver_id")
        booking = Booking(
            requester_name=current_user.full_name,
            requester_id=current_user.id,
            driver_id=int(driver_id) if driver_id else None,
            vehicle_id=vehicle_id,
            start_datetime_planned=start_dt,
            end_datetime_planned=end_dt,
            route_from=route_from,
            route_to=route_to,
            purpose=purpose,
            activity_code=request.form.get("activity_code", "").strip(),
            project_code=request.form.get("project_code", "").strip(),
            status="pending",
        )
        db.session.add(booking)
        db.session.flush()  # assign ID before re-checking

        # Double-check for conflicts after flush to reduce race window
        conflict = check_booking_conflict(vehicle_id, start_dt, end_dt, exclude_booking_id=booking.id)
        if conflict:
            db.session.rollback()
            flash(
                f"This vehicle is already booked between "
                f"{conflict.start_datetime_planned.strftime('%Y-%m-%d %H:%M')} and "
                f"{conflict.end_datetime_planned.strftime('%Y-%m-%d %H:%M')} "
                f"(Booking #{conflict.id} by {conflict.requester_name}).",
                "danger",
            )
            return render_template(
                "bookings/add.html", vehicles=vehicles, drivers=drivers
            )

        db.session.commit()
        log_action("create", "Booking", booking.id, f"Created booking: vehicle={booking.vehicle.registration_number}, route={route_from}→{route_to}")
        flash("Booking request created (status: pending).", "success")

        # ── Notify all admins about the new booking request ────────────
        admins = User.query.filter_by(role="admin", is_active_user=True).all()
        admin_emails = [a.email for a in admins if a.email]
        if admin_emails:
            send_notification(
                subject=f"New Booking Request #{booking.id} – Vehicle Request Tracker",
                recipients=admin_emails,
                body=(
                    f"Hello Admin,\n\n"
                    f"A new vehicle booking request has been submitted.\n\n"
                    f"Booking #: {booking.id}\n"
                    f"Requested by: {booking.requester_name}\n"
                    f"Vehicle: {booking.vehicle.registration_number}\n"
                    f"Route: {booking.route_from} → {booking.route_to}\n"
                    f"Purpose: {booking.purpose}\n"
                    f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                    f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                    f"Please log in to review and approve/reject this request.\n\n"
                    f"– Vehicle Request Tracker"
                ),
            )
        # ── Confirm to the requester that their request was received ──
        if current_user.email:
            send_notification(
                subject=f"Booking Request #{booking.id} Received – Vehicle Request Tracker",
                recipients=[current_user.email],
                body=(
                    f"Hello {current_user.full_name},\n\n"
                    f"Your vehicle booking request has been submitted successfully.\n\n"
                    f"Booking #: {booking.id}\n"
                    f"Vehicle: {booking.vehicle.registration_number}\n"
                    f"Route: {booking.route_from} → {booking.route_to}\n"
                    f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                    f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                    f"Status: PENDING – awaiting admin approval.\n"
                    f"You will receive another email once your request is approved or cancelled.\n\n"
                    f"– Vehicle Request Tracker"
                ),
            )

        return redirect(url_for("booking_list"))

    return render_template("bookings/add.html", vehicles=vehicles, drivers=drivers)


@app.route("/bookings/<int:booking_id>")
@login_required
def booking_detail(booking_id):
    booking = db.get_or_404(Booking, booking_id)
    drivers = User.query.filter_by(role="driver").order_by(User.full_name).all()
    return render_template("bookings/detail.html", booking=booking, drivers=drivers)


@app.route("/bookings/<int:booking_id>/approve", methods=["POST"])
@role_required("admin")
def booking_approve(booking_id):
    """Approve a pending booking – but only if there is NO overlap conflict."""
    booking = db.get_or_404(Booking, booking_id)

    if booking.status != "pending":
        flash("Only pending bookings can be approved.", "warning")
        return redirect(url_for("booking_detail", booking_id=booking.id))

    # ── Conflict check (exclude this booking itself) ──────────────────
    conflict = check_booking_conflict(
        booking.vehicle_id,
        booking.start_datetime_planned,
        booking.end_datetime_planned,
        exclude_booking_id=booking.id,
    )
    if conflict:
        flash(
            f"Cannot approve – this vehicle is already booked between "
            f"{conflict.start_datetime_planned.strftime('%Y-%m-%d %H:%M')} and "
            f"{conflict.end_datetime_planned.strftime('%Y-%m-%d %H:%M')} "
            f"(Booking #{conflict.id} by {conflict.requester_name}).",
            "danger",
        )
        return redirect(url_for("booking_detail", booking_id=booking.id))

    booking.status = "approved"
    db.session.commit()
    log_action("approve", "Booking", booking.id, f"Approved booking #{booking.id} for vehicle {booking.vehicle.registration_number}")
    flash("Booking approved.", "success")

    # ── Send notification email ───────────────────────────────────────
    if booking.requester and booking.requester.email:
        send_notification(
            subject=f"Booking #{booking.id} Approved – Vehicle Request Tracker",
            recipients=[booking.requester.email],
            body=(
                f"Hello {booking.requester_name},\n\n"
                f"Your vehicle booking #{booking.id} has been approved.\n\n"
                f"Vehicle: {booking.vehicle.registration_number}\n"
                f"Route: {booking.route_from} → {booking.route_to}\n"
                f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                f"– Vehicle Request Tracker"
            ),
        )
    # Also notify driver if assigned
    if booking.driver and booking.driver.email:
        send_notification(
            subject=f"You have been assigned to Booking #{booking.id}",
            recipients=[booking.driver.email],
            body=(
                f"Hello {booking.driver.full_name},\n\n"
                f"You have been assigned as driver for booking #{booking.id}.\n\n"
                f"Vehicle: {booking.vehicle.registration_number}\n"
                f"Route: {booking.route_from} → {booking.route_to}\n"
                f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                f"– Vehicle Request Tracker"
            ),
        )

    return redirect(url_for("booking_detail", booking_id=booking.id))


@app.route("/bookings/<int:booking_id>/assign-driver", methods=["POST"])
@role_required("admin")
def booking_assign_driver(booking_id):
    """Assign or change the driver for a booking."""
    booking = db.get_or_404(Booking, booking_id)
    driver_id = request.form.get("driver_id")
    booking.driver_id = int(driver_id) if driver_id else None
    db.session.commit()

    if booking.driver:
        log_action("assign", "Booking", booking.id, f"Assigned driver '{booking.driver.full_name}' to booking #{booking.id}")
        flash(f"Driver assigned: {booking.driver.full_name}.", "success")
        # Notify the driver
        send_notification(
            subject=f"Driver Assignment – Booking #{booking.id}",
            recipients=[booking.driver.email],
            body=(
                f"Hello {booking.driver.full_name},\n\n"
                f"You have been assigned as driver for booking #{booking.id}.\n"
                f"Vehicle: {booking.vehicle.registration_number}\n"
                f"Route: {booking.route_from} → {booking.route_to}\n"
                f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                f"– Vehicle Request Tracker"
            ),
        )
    else:
        flash("Driver removed from booking.", "info")

    return redirect(url_for("booking_detail", booking_id=booking.id))


@app.route("/bookings/<int:booking_id>/cancel", methods=["POST"])
@login_required
def booking_cancel(booking_id):
    booking = db.get_or_404(Booking, booking_id)
    if booking.status in ("pending", "approved"):
        booking.status = "cancelled"
        db.session.commit()
        log_action("cancel", "Booking", booking.id, f"Cancelled booking #{booking.id} for vehicle {booking.vehicle.registration_number}")
        flash("Booking cancelled.", "info")

        # ── Notify the requester that their booking was cancelled ─────
        if booking.requester and booking.requester.email:
            cancelled_by = current_user.full_name
            send_notification(
                subject=f"Booking #{booking.id} Cancelled – Vehicle Request Tracker",
                recipients=[booking.requester.email],
                body=(
                    f"Hello {booking.requester_name},\n\n"
                    f"Your vehicle booking #{booking.id} has been cancelled"
                    f"{' by ' + cancelled_by if cancelled_by != booking.requester_name else ''}.\n\n"
                    f"Vehicle: {booking.vehicle.registration_number}\n"
                    f"Route: {booking.route_from} → {booking.route_to}\n"
                    f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                    f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                    f"If you have questions, please contact the admin.\n\n"
                    f"– Vehicle Request Tracker"
                ),
            )
        # ── Notify assigned driver if any ─────────────────────────────
        if booking.driver and booking.driver.email:
            send_notification(
                subject=f"Booking #{booking.id} Cancelled – Vehicle Request Tracker",
                recipients=[booking.driver.email],
                body=(
                    f"Hello {booking.driver.full_name},\n\n"
                    f"Booking #{booking.id} you were assigned to has been cancelled.\n\n"
                    f"Vehicle: {booking.vehicle.registration_number}\n"
                    f"Route: {booking.route_from} → {booking.route_to}\n"
                    f"From: {booking.start_datetime_planned.strftime('%d %b %Y %H:%M')}\n"
                    f"To: {booking.end_datetime_planned.strftime('%d %b %Y %H:%M')}\n\n"
                    f"– Vehicle Request Tracker"
                ),
            )
    else:
        flash("This booking cannot be cancelled.", "warning")
    return redirect(url_for("booking_detail", booking_id=booking.id))


@app.route("/bookings/<int:booking_id>/delete", methods=["POST"])
@role_required("admin")
def booking_delete(booking_id):
    booking = db.get_or_404(Booking, booking_id)
    vehicle_reg = booking.vehicle.registration_number
    # Delete associated trip first
    if booking.trip:
        db.session.delete(booking.trip)
    db.session.delete(booking)
    db.session.commit()
    log_action("delete", "Booking", booking_id, f"Deleted booking #{booking_id} ({vehicle_reg})")
    flash(f"Booking #{booking_id} ({vehicle_reg}) has been deleted.", "success")
    return redirect(url_for("booking_list"))


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  TRIPS                                                                  ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/bookings/<int:booking_id>/trip/start", methods=["GET", "POST"])
@login_required
def trip_start(booking_id):
    """Record the actual start of a trip (for an approved booking)."""
    booking = db.get_or_404(Booking, booking_id)

    if booking.status != "approved":
        flash("Only approved bookings can start a trip.", "warning")
        return redirect(url_for("booking_detail", booking_id=booking.id))

    if booking.trip is not None:
        flash("A trip has already been started for this booking.", "warning")
        return redirect(url_for("booking_detail", booking_id=booking.id))

    if request.method == "POST":
        errors = []

        start_raw = request.form.get("start_actual_datetime", "")
        odo_raw = request.form.get("odometer_start", "")

        start_dt = None
        if not start_raw:
            errors.append("Start date/time is required.")
        else:
            try:
                start_dt = datetime.fromisoformat(start_raw)
            except ValueError:
                errors.append("Invalid start date/time format.")

        odometer_start = None
        if not odo_raw:
            errors.append("Odometer reading is required.")
        else:
            try:
                odometer_start = int(odo_raw)
                if odometer_start < 0:
                    errors.append("Odometer reading cannot be negative.")
            except ValueError:
                errors.append("Odometer reading must be a whole number.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("trips/start.html", booking=booking)

        trip = Trip(
            booking_id=booking.id,
            start_actual_datetime=start_dt,
            odometer_start=odometer_start,
        )
        # Mark vehicle as in use
        booking.vehicle.status = "in_use"
        db.session.add(trip)
        db.session.commit()
        log_action("create", "Trip", trip.id, f"Started trip for booking #{booking.id}, odometer={odometer_start}")
        flash("Trip started – vehicle marked as in use.", "success")
        return redirect(url_for("booking_detail", booking_id=booking.id))

    return render_template("trips/start.html", booking=booking)


@app.route("/bookings/<int:booking_id>/trip/end", methods=["GET", "POST"])
@login_required
def trip_end(booking_id):
    """Record the actual end of a trip and mark the booking as completed."""
    booking = db.get_or_404(Booking, booking_id)
    trip = booking.trip

    if trip is None or trip.end_actual_datetime is not None:
        flash("No active trip found for this booking.", "warning")
        return redirect(url_for("booking_detail", booking_id=booking.id))

    if request.method == "POST":
        errors = []

        end_raw = request.form.get("end_actual_datetime", "")
        odo_end_raw = request.form.get("odometer_end", "")
        fuel_raw = request.form.get("fuel_used", "").strip()
        cost_raw = request.form.get("fuel_cost", "").strip()

        end_dt = None
        if not end_raw:
            errors.append("End date/time is required.")
        else:
            try:
                end_dt = datetime.fromisoformat(end_raw)
            except ValueError:
                errors.append("Invalid end date/time format.")

        odometer_end = None
        if not odo_end_raw:
            errors.append("Odometer reading is required.")
        else:
            try:
                odometer_end = int(odo_end_raw)
            except ValueError:
                errors.append("Odometer reading must be a whole number.")

        # Cross-field checks
        if end_dt and trip.start_actual_datetime and end_dt <= trip.start_actual_datetime:
            errors.append("End date/time must be after the trip start time.")

        if odometer_end is not None and odometer_end < trip.odometer_start:
            errors.append(
                f"End odometer ({odometer_end}) cannot be less than start "
                f"odometer ({trip.odometer_start})."
            )

        # Optional numeric fields
        fuel_used = None
        if fuel_raw:
            try:
                fuel_used = float(fuel_raw)
                if fuel_used < 0:
                    errors.append("Fuel used cannot be negative.")
            except ValueError:
                errors.append("Fuel used must be a number.")

        fuel_cost = None
        if cost_raw:
            try:
                fuel_cost = float(cost_raw)
                if fuel_cost < 0:
                    errors.append("Fuel cost cannot be negative.")
            except ValueError:
                errors.append("Fuel cost must be a number.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("trips/end.html", booking=booking, trip=trip)

        trip.end_actual_datetime = end_dt
        trip.odometer_end = odometer_end
        trip.distance = odometer_end - trip.odometer_start
        trip.fuel_used = fuel_used
        trip.fuel_cost = fuel_cost
        trip.remarks = request.form.get("remarks", "").strip() or None

        # Mark booking completed and vehicle available again
        booking.status = "completed"
        booking.vehicle.status = "available"
        db.session.commit()
        log_action("complete", "Trip", trip.id, f"Ended trip for booking #{booking.id}, distance={trip.distance} km")
        flash(
            f"Trip ended – distance: {trip.distance} km. Booking marked as completed.",
            "success",
        )
        return redirect(url_for("booking_detail", booking_id=booking.id))

    return render_template("trips/end.html", booking=booking, trip=trip)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  MAINTENANCE                                                            ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/maintenance")
@login_required
def maintenance_list():
    status_filter = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)
    query = MaintenanceRecord.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    pagination = query.order_by(MaintenanceRecord.scheduled_date.desc()).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    return render_template(
        "maintenance/list.html",
        records=pagination.items,
        pagination=pagination,
        current_status=status_filter,
    )


@app.route("/maintenance/add", methods=["GET", "POST"])
@role_required("admin")
def maintenance_add():
    vehicles = Vehicle.query.order_by(Vehicle.registration_number).all()
    if request.method == "POST":
        errors = []

        try:
            vehicle_id = int(request.form.get("vehicle_id", 0))
        except (ValueError, TypeError):
            vehicle_id = 0
        if not vehicle_id:
            errors.append("Please select a vehicle.")

        mtype = request.form.get("maintenance_type", "")
        if mtype not in ("routine", "repair", "inspection", "tyre", "other"):
            errors.append("Invalid maintenance type.")

        description = request.form.get("description", "").strip()
        if not description:
            errors.append("Description is required.")

        sched_raw = request.form.get("scheduled_date", "")
        sched_date = None
        if not sched_raw:
            errors.append("Scheduled date is required.")
        else:
            try:
                sched_date = datetime.strptime(sched_raw, "%Y-%m-%d").date()
            except ValueError:
                errors.append("Invalid date format.")

        cost = None
        cost_raw = request.form.get("cost", "").strip()
        if cost_raw:
            try:
                cost = float(cost_raw)
                if cost < 0:
                    errors.append("Cost cannot be negative.")
            except ValueError:
                errors.append("Cost must be a number.")

        if errors:
            for e in errors:
                flash(e, "danger")
            return render_template("maintenance/add.html", vehicles=vehicles)

        rec = MaintenanceRecord(
            vehicle_id=vehicle_id,
            maintenance_type=mtype,
            description=description,
            scheduled_date=sched_date,
            cost=cost,
            status="scheduled",
            created_by_id=current_user.id,
        )
        db.session.add(rec)
        db.session.commit()

        # Optionally set vehicle to maintenance status
        if request.form.get("set_maintenance"):
            rec.vehicle.status = "maintenance"
            db.session.commit()

        log_action("create", "MaintenanceRecord", rec.id, f"Created maintenance ({mtype}) for vehicle {rec.vehicle.registration_number}")
        flash("Maintenance record created.", "success")
        return redirect(url_for("maintenance_list"))
    return render_template("maintenance/add.html", vehicles=vehicles)


@app.route("/maintenance/<int:rec_id>/complete", methods=["POST"])
@role_required("admin")
def maintenance_complete(rec_id):
    rec = db.get_or_404(MaintenanceRecord, rec_id)
    rec.status = "completed"
    rec.completed_date = datetime.now(timezone.utc).date()
    if request.form.get("cost"):
        rec.cost = float(request.form["cost"])
    # Set vehicle back to available
    rec.vehicle.status = "available"
    db.session.commit()
    log_action("complete", "MaintenanceRecord", rec.id, f"Completed maintenance for vehicle {rec.vehicle.registration_number}")
    flash("Maintenance marked as completed. Vehicle is now available.", "success")
    return redirect(url_for("maintenance_list"))


@app.route("/maintenance/<int:rec_id>/cancel", methods=["POST"])
@role_required("admin")
def maintenance_cancel(rec_id):
    rec = db.get_or_404(MaintenanceRecord, rec_id)
    rec.status = "cancelled"
    if rec.vehicle.status == "maintenance":
        rec.vehicle.status = "available"
    db.session.commit()
    log_action("cancel", "MaintenanceRecord", rec.id, f"Cancelled maintenance for vehicle {rec.vehicle.registration_number}")
    flash("Maintenance record cancelled.", "info")
    return redirect(url_for("maintenance_list"))


@app.route("/maintenance/<int:rec_id>/delete", methods=["POST"])
@role_required("admin")
def maintenance_delete(rec_id):
    rec = db.get_or_404(MaintenanceRecord, rec_id)
    vehicle_reg = rec.vehicle.registration_number
    db.session.delete(rec)
    db.session.commit()
    log_action("delete", "MaintenanceRecord", rec_id, f"Deleted maintenance record for vehicle '{vehicle_reg}'")
    flash(f"Maintenance record for '{vehicle_reg}' has been deleted.", "success")
    return redirect(url_for("maintenance_list"))


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  REPORTS                                                                ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/reports/vehicle", methods=["GET"])
@login_required
def vehicle_report():
    vehicles = Vehicle.query.order_by(Vehicle.registration_number).all()

    trips = []
    total_distance = 0
    total_fuel_cost = 0
    selected_vehicle_id = request.args.get("vehicle_id", type=int)
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    if selected_vehicle_id and date_from and date_to:
        dt_from = datetime.fromisoformat(date_from)
        dt_to = datetime.fromisoformat(date_to + "T23:59:59")

        trips = (
            Trip.query.join(Booking)
            .filter(
                Booking.vehicle_id == selected_vehicle_id,
                Trip.start_actual_datetime >= dt_from,
                Trip.start_actual_datetime <= dt_to,
                Trip.end_actual_datetime.isnot(None),
            )
            .order_by(Trip.start_actual_datetime)
            .all()
        )
        total_distance = sum(t.distance or 0 for t in trips)
        total_fuel_cost = sum(t.fuel_cost or 0 for t in trips)

    return render_template(
        "reports/vehicle_report.html",
        vehicles=vehicles,
        trips=trips,
        total_distance=total_distance,
        total_fuel_cost=total_fuel_cost,
        selected_vehicle_id=selected_vehicle_id,
        date_from=date_from,
        date_to=date_to,
    )


# ── Budget report per project code ───────────────────────────────────────────


@app.route("/reports/budget")
@login_required
def budget_report():
    """Fuel cost summary grouped by project_code."""
    rows = (
        db.session.query(
            Booking.project_code,
            db.func.count(Trip.id).label("trip_count"),
            db.func.sum(Trip.distance).label("total_distance"),
            db.func.sum(Trip.fuel_used).label("total_fuel"),
            db.func.sum(Trip.fuel_cost).label("total_cost"),
        )
        .join(Trip, Trip.booking_id == Booking.id)
        .filter(Trip.end_actual_datetime.isnot(None))
        .group_by(Booking.project_code)
        .all()
    )
    return render_template("reports/budget_report.html", rows=rows)


# ── Excel export ─────────────────────────────────────────────────────────────


@app.route("/reports/vehicle/export")
@login_required
def vehicle_report_export():
    """Export trip report to an Excel (.xlsx) file."""
    selected_vehicle_id = request.args.get("vehicle_id", type=int)
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")

    if not (selected_vehicle_id and date_from and date_to):
        flash("Please select a vehicle and date range first.", "warning")
        return redirect(url_for("vehicle_report"))

    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to + "T23:59:59")

    vehicle = db.get_or_404(Vehicle, selected_vehicle_id)

    trips = (
        Trip.query.join(Booking)
        .filter(
            Booking.vehicle_id == selected_vehicle_id,
            Trip.start_actual_datetime >= dt_from,
            Trip.start_actual_datetime <= dt_to,
            Trip.end_actual_datetime.isnot(None),
        )
        .order_by(Trip.start_actual_datetime)
        .all()
    )

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trip Report"

    # Title row
    header_font = Font(bold=True, size=14)
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Trip Report – {vehicle.registration_number} ({vehicle.make} {vehicle.model})"
    ws["A1"].font = header_font

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Period: {date_from} to {date_to}"
    ws["A2"].font = Font(size=11, italic=True)

    # Column headers
    headers = [
        "Trip #", "Booking #", "Requester", "Driver", "Route",
        "Start", "End", "Distance (km)", "Fuel (L)", "Fuel Cost",
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, t in enumerate(trips, 5):
        ws.cell(row=i, column=1, value=t.id)
        ws.cell(row=i, column=2, value=t.booking.id)
        ws.cell(row=i, column=3, value=t.booking.requester_name)
        ws.cell(
            row=i, column=4,
            value=t.booking.driver.full_name if t.booking.driver else "–",
        )
        ws.cell(row=i, column=5, value=f"{t.booking.route_from} → {t.booking.route_to}")
        ws.cell(row=i, column=6, value=t.start_actual_datetime.strftime("%d %b %Y %H:%M"))
        ws.cell(row=i, column=7, value=t.end_actual_datetime.strftime("%d %b %Y %H:%M"))
        ws.cell(row=i, column=8, value=t.distance)
        ws.cell(row=i, column=9, value=t.fuel_used or "")
        ws.cell(row=i, column=10, value=t.fuel_cost or "")

    # Totals row
    total_row = len(trips) + 5
    ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=sum(t.distance or 0 for t in trips)).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=sum(t.fuel_used or 0 for t in trips)).font = Font(bold=True)
    ws.cell(row=total_row, column=10, value=sum(t.fuel_cost or 0 for t in trips)).font = Font(bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=12)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"trip_report_{vehicle.registration_number}_{date_from}_to_{date_to}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  CALENDAR (JSON API + page)                                             ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/calendar")
@login_required
def calendar_view():
    return render_template("calendar.html")


@app.route("/api/bookings")
@login_required
def api_bookings():
    """Return bookings as JSON events for FullCalendar."""
    bookings = Booking.query.filter(
        Booking.status.in_(["pending", "approved"])
    ).all()

    colour_map = {
        "pending": "#ffc107",   # yellow
        "approved": "#198754",  # green
    }

    events = []
    for b in bookings:
        # Sanitise user-provided fields for defence-in-depth
        safe_from = (b.route_from or "").replace("<", "&lt;").replace(">", "&gt;")
        safe_to = (b.route_to or "").replace("<", "&lt;").replace(">", "&gt;")
        events.append(
            {
                "id": b.id,
                "title": f"{b.vehicle.registration_number} – {safe_from}→{safe_to}",
                "start": b.start_datetime_planned.isoformat(),
                "end": b.end_datetime_planned.isoformat(),
                "url": url_for("booking_detail", booking_id=b.id),
                "color": colour_map.get(b.status, "#6c757d"),
            }
        )
    return jsonify(events)


# ╔═══════════════════════════════════════════════════════════════════════════╗
# ║  AUDIT LOG                                                              ║
# ╚═══════════════════════════════════════════════════════════════════════════╝


@app.route("/audit-log")
@role_required("admin")
def audit_log():
    """Show the audit trail of all create / edit / delete actions."""
    entity_filter = request.args.get("entity", "")
    action_filter = request.args.get("action", "")
    page = request.args.get("page", 1, type=int)

    query = AuditLog.query
    if entity_filter:
        query = query.filter_by(entity_type=entity_filter)
    if action_filter:
        query = query.filter_by(action=action_filter)

    pagination = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    return render_template(
        "audit_log.html",
        logs=pagination.items,
        pagination=pagination,
        current_entity=entity_filter,
        current_action=action_filter,
    )


# ── Run ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
